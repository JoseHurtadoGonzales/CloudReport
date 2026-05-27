"""
HTML → XLSX worker. Parses tables from rendered HTML with pandas and writes
them to an Excel workbook. Honors options:

    sheetNames:  [str, ...]   names for each table found
    headerBold:  bool         bold first row (default true)
    autofit:     bool         autofit column widths (default true)
    zebra:       bool         alternate-row fill
    exportPdf:   bool         post-convert to PDF via LibreOffice
"""
from __future__ import annotations

import io
import os
import sys

sys.path.insert(0, "/app/shared")

from protocol import Job, Reply, S3, consume  # noqa: E402

import pandas as pd  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

try:
    from libreoffice import convert_to_pdf  # noqa: E402
    HAS_SOFFICE = True
except Exception:
    HAS_SOFFICE = False


s3 = S3()


def autofit_columns(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def render(job: Job) -> Reply:
    data = job.data or {}
    html = data.get("html") or ""
    options = data.get("options") or {}

    header_bold = bool(options.get("headerBold", True))
    autofit = bool(options.get("autofit", True))
    zebra = bool(options.get("zebra", False))
    export_pdf = bool(options.get("exportPdf", False))
    frozen_rows = int(options.get("frozenRows", 1))

    if not html.strip():
        return Reply(id=job.id, status="error", error="html-to-xlsx: empty html")

    try:
        tables = pd.read_html(io.StringIO(html))
    except ValueError:
        tables = []

    wb = Workbook()
    if not tables:
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=1, column=1, value=html)
    else:
        wb.remove(wb.active)
        names = options.get("sheetNames") or []
        zebra_fill = PatternFill("solid", fgColor="F5F7FA")
        for idx, df in enumerate(tables):
            sheet_name = (
                names[idx] if idx < len(names) and names[idx]
                else f"Sheet{idx + 1}"
            )[:31]
            ws = wb.create_sheet(sheet_name)
            # Header
            for c, col in enumerate(df.columns.tolist(), start=1):
                cell = ws.cell(row=1, column=c, value=str(col))
                if header_bold:
                    cell.font = Font(bold=True)
            # Body
            for r, row in enumerate(df.itertuples(index=False), start=2):
                for c, val in enumerate(row, start=1):
                    if pd.isna(val):
                        val = None
                    cell = ws.cell(row=r, column=c, value=val)
                    if zebra and r % 2 == 0:
                        cell.fill = zebra_fill
            if autofit:
                autofit_columns(ws)
            if frozen_rows > 0:
                ws.freeze_panes = ws.cell(row=frozen_rows + 1, column=1)

    out = io.BytesIO()
    wb.save(out)
    out_bytes = out.getvalue()

    if export_pdf:
        if not HAS_SOFFICE:
            return Reply(id=job.id, status="error",
                         error="exportPdf requires LibreOffice in the worker image")
        try:
            out_bytes = convert_to_pdf(out_bytes, "xlsx")
        except Exception as e:
            return Reply(id=job.id, status="error",
                         error=f"xlsx→pdf failed: {e}")
        mime = "application/pdf"
        prefix = "outputs/xlsx-pdf"
    else:
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        prefix = "outputs/xlsx"

    key = s3.put(prefix, out_bytes, mime)
    return Reply(id=job.id, status="success", outputBlob=key, mimeType=mime)


def main() -> None:
    stream = os.environ.get("WORKER_STREAM", "renders:html-to-xlsx")
    group = os.environ.get("WORKER_GROUP", "html-to-xlsx-workers")
    consume(stream, group, render)


if __name__ == "__main__":
    main()
